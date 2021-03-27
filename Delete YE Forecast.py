import os
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font


folder_path = r'C:\Users\jpacker\Desktop\Center Reports'
save_path = r'C:\Users\jpacker\Desktop\Center Reports'

for file in os.listdir(folder_path):
	if(file[-4:] == "xlsx"):
		wb = openpyxl.load_workbook(folder_path + "\\" + file)
		
		sheet1 = wb._sheets[0]
		sheet2 = wb._sheets[1]
		sheet3 = wb._sheets[2]
		
		#modifications to sheet 1
		sheet1.delete_cols(19, 4)
		sheet1['R35'].value = None
		medium = Side(border_style="medium", color="000000")
		thin = Side(border_style="thin", color="000000")
		for i in range(34,40):
			if i == 34:
				sheet1['R'+str(i)].border = Border(right=thin,top=thin)
			elif i == 39:
				sheet1['R'+str(i)].border = Border(right=thin,bottom=thin)
			else:
				sheet1['R'+str(i)].border = Border(right=thin)
		sheet1.column_dimensions['S'].hidden= False
		r = 0
		for j in range(30,5000):
			if sheet1['D'+str(j)].value == "TOTALS":
				r = j
				break
		for k in range(r+1,r+5):
			if k == r+1:
				sheet1['R'+str(k)].border = Border(top=thin,right=medium)
			elif k == r+4:
				sheet1['R'+str(k)].border = Border(right=medium,bottom=medium)
			else:
				sheet1['R'+str(k)].border = Border(right=medium)
				
		
		#modifications to sheet2
		sheet2.row_dimensions[32].hidden = False
		for i in range(25,29):
			sheet2.unmerge_cells('X'+str(i)+':'+'Y'+str(i))
			sheet2.unmerge_cells('U'+str(i)+':'+'V'+str(i))
			sheet2['T'+str(i)].value = None
		sheet2.unmerge_cells('U'+str(30)+':'+'X'+str(30))	
		for col in ['U','V','W','X']:
			sheet2.column_dimensions[col].hidden= True
		nothing = Side(border_style=None)
		for i in range(25,29):
			sheet2['Y'+str(i)].border = Border(right=nothing,left=nothing,top=nothing,bottom=nothing)
		sheet2.column_dimensions['Y'].hidden= False
		sheet2.column_dimensions['Y'].width = 9
		
		#modifications to sheet3
		sheet3.row_dimensions[32].hidden = False
		for i in range(25,29):
			sheet3.unmerge_cells('X'+str(i)+':'+'Y'+str(i))
			sheet3.unmerge_cells('U'+str(i)+':'+'V'+str(i))
			sheet3['T'+str(i)].value = None
		sheet3.unmerge_cells('U'+str(30)+':'+'X'+str(30))	
		for col in ['U','V','W','X']:
			sheet3.column_dimensions[col].hidden= True
		nothing = Side(border_style=None)
		for i in range(25,29):
			sheet3['Y'+str(i)].border = Border(right=nothing,left=nothing,top=nothing,bottom=nothing)
		sheet3.column_dimensions['Y'].hidden= False
		sheet3.column_dimensions['Y'].width = 9
	
		
		wb.save(folder_path + "\\" + file)